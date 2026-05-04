# player/views_dashboard.py
"""
Vistas del Dashboard de Estadísticas.

Incluye:
- DashboardIndexView: Vista principal del dashboard con estadísticas generales
- TeamStatsView: Estadísticas detalladas del equipo
- MatchStatsView: Estadísticas detalladas de un partido específico
- CompareView: Comparador de partidos
"""
import csv
import io
import json
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.views.generic import TemplateView, View
from openpyxl import load_workbook

from .models import Match, CoachTournamentTeamParticipation, Team, Profile, GpsMetric
from .services.stats_service import StatsService


class DashboardAccessMixin(LoginRequiredMixin):
    """Mixin para validar acceso y resolver equipos asignados."""

    def get_user_teams(self):
        user = self.request.user
        if user.is_staff:
            return list(Team.objects.order_by('name').values_list('name', flat=True))

        teams = set()

        profile = Profile.objects.filter(user=user).select_related('team').first()
        if profile and profile.team and profile.team.name:
            teams.add(profile.team.name)

        participations = CoachTournamentTeamParticipation.objects.filter(
            user=user,
            active=True,
        ).select_related('team')
        for participation in participations:
            if participation.team and participation.team.name:
                teams.add(participation.team.name)

        return sorted(teams)


class DashboardIndexView(DashboardAccessMixin, TemplateView):
    """Vista principal del dashboard."""

    template_name = 'player/dashboard/index.html'

    def _get_next_match_context(self, teams_to_check, effective_team):
        """Busca el próximo partido para los equipos del entrenador y videos del rival."""
        if not teams_to_check:
            return {'next_match': None}

        today = timezone.localdate()
        teams_upper = set(t.strip().upper() for t in teams_to_check)

        # Incluir alias de cada equipo para que la búsqueda funcione
        # independientemente de como quedó guardado el nombre en Match
        alias_filter = Q()
        for t in teams_to_check:
            alias_filter |= Q(name__iexact=t)
        for obj in Team.objects.filter(alias_filter).exclude(alias__isnull=True).exclude(alias=''):
            teams_upper.add(obj.alias.strip().upper())

        next_match = (
            Match.objects
            .filter(
                Q(home_team__in=teams_upper) | Q(away_team__in=teams_upper),
                match_date__gte=today,
            )
            .order_by('match_date', 'match_time')
            .select_related('tournament')
            .first()
        )

        if not next_match:
            return {'next_match': None}

        # Determinar cuál es nuestro equipo y cuál es el rival
        if next_match.home_team in teams_upper:
            my_team = next_match.home_team
            rival = next_match.away_team
        else:
            my_team = next_match.away_team
            rival = next_match.home_team

        # Videos del rival: ya no se muestran en el dashboard (pendiente vista de estadísticas rivales)

        return {
            'next_match': next_match,
            'next_match_my_team': my_team,
            'next_match_rival': rival,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        selected_seasons = self.request.GET.getlist('season', [])
        selected_tournaments = [t for t in self.request.GET.getlist('tournament', []) if t]
        selected_team = self.request.GET.get('team')

        user_teams = self.get_user_teams()
        is_admin = self.request.user.is_staff

        if is_admin and not selected_team:
            context['user_teams'] = user_teams
            context['selected_team'] = None
            context['is_admin'] = True
            context['available_seasons'] = []
            context['available_tournaments'] = []
            context['selected_seasons'] = selected_seasons
            context['selected_tournaments'] = selected_tournaments
            context['needs_team_selection'] = True
            context['no_teams_assigned'] = False
            return context

        if not is_admin and not user_teams:
            context['user_teams'] = []
            context['selected_team'] = None
            context['is_admin'] = is_admin
            context['available_seasons'] = []
            context['available_tournaments'] = []
            context['selected_seasons'] = selected_seasons
            context['selected_tournaments'] = selected_tournaments
            context['needs_team_selection'] = False
            context['no_teams_assigned'] = True
            return context

        effective_team = selected_team or (None if is_admin else (user_teams[0] if user_teams else None))

        stats_service = StatsService(
            user=self.request.user,
            team_name=effective_team,
            seasons=selected_seasons if selected_seasons else None,
            tournaments=selected_tournaments if selected_tournaments else None,
        )

        available_seasons = stats_service.get_available_seasons()
        available_tournaments = stats_service.get_available_tournaments()

        context['summary'] = stats_service.get_summary_stats()
        context['recent_matches'] = stats_service.get_recent_matches(limit=5)
        context['plays_distribution'] = stats_service.get_plays_distribution()
        context['trend_data'] = stats_service.get_trend_data(last_n_matches=10)
        context['zone_data'] = stats_service.get_zone_heatmap_data()
        context['season_aggregates'] = stats_service.get_season_aggregates()

        context['available_seasons'] = available_seasons
        context['available_tournaments'] = available_tournaments
        context['selected_seasons'] = selected_seasons
        context['selected_tournaments'] = selected_tournaments
        context['user_teams'] = user_teams
        context['selected_team'] = effective_team
        context['is_admin'] = is_admin
        context['needs_team_selection'] = False
        context['no_teams_assigned'] = False

        context['trend_data_json'] = json.dumps(context['trend_data'])
        context['plays_distribution_json'] = json.dumps(context['plays_distribution'])
        context['zone_data_json'] = json.dumps(context['zone_data'])

        # Próximo partido y videos del rival
        teams_to_check = [effective_team] if effective_team else user_teams
        context.update(self._get_next_match_context(teams_to_check, effective_team))

        return context


class TeamStatsView(DashboardAccessMixin, TemplateView):
    """Vista de estadísticas detalladas del equipo."""
    
    template_name = 'player/dashboard/team_stats.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        selected_seasons = self.request.GET.getlist('season', [])
        selected_tournaments = [t for t in self.request.GET.getlist('tournament', []) if t]
        selected_team = self.request.GET.get('team')
        user_teams = self.get_user_teams()
        is_admin = self.request.user.is_staff
        effective_team = selected_team or (None if is_admin else (user_teams[0] if user_teams else None))
        
        stats_service = StatsService(
            user=self.request.user,
            team_name=effective_team,
            seasons=selected_seasons if selected_seasons else None,
            tournaments=selected_tournaments if selected_tournaments else None
        )
        
        context['summary'] = stats_service.get_summary_stats()
        context['plays_distribution'] = stats_service.get_plays_distribution()
        context['trend_data'] = stats_service.get_trend_data(last_n_matches=20)
        context['zone_data'] = stats_service.get_zone_heatmap_data()
        context['season_aggregates'] = stats_service.get_season_aggregates()
        
        context['available_seasons'] = stats_service.get_available_seasons()
        context['available_tournaments'] = stats_service.get_available_tournaments()
        context['selected_seasons'] = selected_seasons
        context['selected_tournaments'] = selected_tournaments
        context['user_teams'] = user_teams
        context['selected_team'] = effective_team
        
        # JSON para gráficos
        context['trend_data_json'] = json.dumps(context['trend_data'])
        context['plays_distribution_json'] = json.dumps(context['plays_distribution'])
        context['zone_data_json'] = json.dumps(context['zone_data'])
        
        return context


class MatchStatsView(DashboardAccessMixin, TemplateView):
    """Vista de estadísticas de un partido específico."""
    
    template_name = 'player/dashboard/match_stats.html'
    
    EXPECTED_COLUMNS = {
        'name': ['name', 'player', 'jugador'],
        'total_distance': ['total distance'],
        'metres_per_minute': ['metres per minute', 'meters per minute', 'm/min'],
        'high_speed_running': ['high speed running', 'high speed running absolute', 'high speed running(absolute)'],
        'accelerations': ['accelerations'],
        'decelerations': ['decelerations'],
        'hml_distance': ['hml distance'],
        'sprints': ['sprints'],
        'sprint_distance': ['sprint distance'],
    }

    def post(self, request, *args, **kwargs):
        match_id = kwargs.get('pk')
        match = get_object_or_404(Match, pk=match_id)

        if not request.user.is_staff:
            messages.error(request, "Solo un usuario administrador puede cargar métricas GPS.")
            return redirect(self._redirect_url(match))

        # Borrado explícito de métricas sin requerir archivo de carga
        if request.POST.get('clear_gps'):
            GpsMetric.objects.filter(match=match).delete()
            messages.success(request, "Se borraron todas las métricas GPS de este partido.")
            return redirect(self._redirect_url(match))

        upload = request.FILES.get('gps_file')
        if not upload:
            messages.error(request, "Debes seleccionar un archivo CSV o XLSX.")
            return redirect(self._redirect_url(match))

        try:
            rows = self._parse_gps_file(upload)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(self._redirect_url(match))

        if not rows:
            messages.warning(request, "El archivo no contiene filas de datos válidas.")
            return redirect(self._redirect_url(match))

        objs = []
        for row in rows:
            objs.append(GpsMetric(
                match=match,
                name=row.get('name', ''),
                total_distance=row.get('total_distance'),
                metres_per_minute=row.get('metres_per_minute'),
                high_speed_running=row.get('high_speed_running'),
                accelerations=row.get('accelerations'),
                decelerations=row.get('decelerations'),
                hml_distance=row.get('hml_distance'),
                sprints=row.get('sprints'),
                sprint_distance=row.get('sprint_distance'),
            ))

        with transaction.atomic():
            GpsMetric.objects.filter(match=match).delete()
            GpsMetric.objects.bulk_create(objs)

        messages.success(request, f"Se cargaron {len(objs)} métricas GPS para el partido.")
        return redirect(self._redirect_url(match))

    def _redirect_url(self, match):
        qs = self.request.GET.urlencode()
        base = reverse('player:dashboard_match', args=[match.pk])
        return f"{base}?{qs}" if qs else base

    def _normalize(self, header: str) -> str:
        if header is None:
            return ''
        s = str(header).strip().lower()
        for ch in [' ', '\t', '\n', '\r', '-', '_']:
            s = s.replace(ch, '')
        s = s.replace('(', '').replace(')', '')
        return s

    def _match_key(self, header: str):
        norm = self._normalize(header)
        for key, aliases in self.EXPECTED_COLUMNS.items():
            for alias in aliases:
                if norm == self._normalize(alias):
                    return key
        return None

    def _to_decimal(self, val):
        if val is None:
            return None
        s = str(val).strip()
        if s == '':
            return None
        s = s.replace(',', '.')
        try:
            return Decimal(s)
        except Exception:
            return None

    def _to_int(self, val):
        if val is None:
            return None
        s = str(val).strip()
        if s == '':
            return None
        try:
            return int(float(s))
        except Exception:
            return None

    def _parse_gps_file(self, uploaded_file):
        name = (uploaded_file.name or '').lower()
        if name.endswith('.csv'):
            text = uploaded_file.read()
            decoded = None
            for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                try:
                    decoded = text.decode(enc)
                    break
                except Exception:
                    decoded = None
            if decoded is None:
                raise ValueError('No se pudo decodificar el CSV (prueba UTF-8).')
            reader = csv.DictReader(io.StringIO(decoded))
            headers = reader.fieldnames or []
            mapping = {h: self._match_key(h) for h in headers}
            if 'name' not in mapping.values():
                raise ValueError('El archivo debe incluir la columna Name.')
            rows = []
            for row in reader:
                data = {}
                for raw_key, value in row.items():
                    key = mapping.get(raw_key)
                    if not key:
                        continue
                    if key == 'name':
                        data['name'] = (value or '').strip()
                    elif key in {'accelerations', 'decelerations', 'sprints'}:
                        data[key] = self._to_int(value)
                    else:
                        data[key] = self._to_decimal(value)
                if data.get('name'):
                    rows.append(data)
            return rows

        if name.endswith('.xlsx') or name.endswith('.xlsm'):
            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            rows_iter = list(ws.iter_rows(values_only=True))
            if not rows_iter:
                return []
            headers = rows_iter[0]
            mapping = {idx: self._match_key(h) for idx, h in enumerate(headers)}
            if 'name' not in mapping.values():
                raise ValueError('El archivo debe incluir la columna Name.')
            rows = []
            for raw in rows_iter[1:]:
                data = {}
                for idx, value in enumerate(raw):
                    key = mapping.get(idx)
                    if not key:
                        continue
                    if key == 'name':
                        data['name'] = (str(value or '').strip())
                    elif key in {'accelerations', 'decelerations', 'sprints'}:
                        data[key] = self._to_int(value)
                    else:
                        data[key] = self._to_decimal(value)
                if data.get('name'):
                    rows.append(data)
            return rows

        raise ValueError('Formato no soportado. Subí un CSV o XLSX.')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        match_id = kwargs.get('pk')
        match = get_object_or_404(Match, pk=match_id)
        selected_team = self.request.GET.get('team')
        
        stats_service = StatsService(user=self.request.user, team_name=selected_team)
        
        context['match'] = match
        context['match_stats'] = stats_service.get_match_detailed_stats(match_id)
        context['selected_team'] = selected_team
        context['gps_metrics'] = list(GpsMetric.objects.filter(match=match).order_by('name'))
        
        # JSON para gráficos
        context['match_stats_json'] = json.dumps(context['match_stats'], default=str)
        
        return context


class CompareMatchesView(DashboardAccessMixin, TemplateView):
    """Vista para comparar múltiples partidos."""

    template_name = 'player/dashboard/compare.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # IDs de partidos a comparar (pueden venir por GET o POST)
        match_ids = self.request.GET.getlist('match_id', [])
        match_ids = [int(mid) for mid in match_ids if mid.isdigit()]

        selected_seasons = self.request.GET.getlist('season', [])
        selected_team = self.request.GET.get('team', None)

        stats_service = StatsService(
            user=self.request.user,
            team_name=selected_team,
            seasons=selected_seasons if selected_seasons else None
        )

        # Obtener partidos disponibles para comparar
        context['available_matches'] = stats_service.get_recent_matches(limit=20)
        context['selected_match_ids'] = match_ids

        # Si hay partidos seleccionados, comparar
        if match_ids:
            context['comparison'] = stats_service.compare_matches(match_ids)
            context['comparison_json'] = json.dumps(context['comparison'], default=str)
        else:
            context['comparison'] = None
            context['comparison_json'] = 'null'

        context['available_seasons'] = stats_service.get_available_seasons()
        context['selected_seasons'] = selected_seasons
        context['user_teams'] = self.get_user_teams()
        context['selected_team'] = selected_team

        return context


class RivalCompareView(DashboardAccessMixin, TemplateView):
    """Comparación estadística entre el equipo del usuario y el próximo rival."""

    template_name = 'player/dashboard/rival_compare.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        selected_seasons = self.request.GET.getlist('season', [])
        selected_tournaments = [t for t in self.request.GET.getlist('tournament', []) if t]
        rival_tournaments = [t for t in self.request.GET.getlist('rival_tournament', []) if t]
        selected_team = self.request.GET.get('team')
        rival_override = self.request.GET.get('rival')  # permite elegir rival manualmente

        user_teams = self.get_user_teams()
        is_admin = self.request.user.is_staff
        effective_team = selected_team or (None if is_admin else (user_teams[0] if user_teams else None))

        context['user_teams'] = user_teams
        context['selected_team'] = effective_team
        context['is_admin'] = is_admin
        context['selected_seasons'] = selected_seasons
        context['selected_tournaments'] = selected_tournaments
        context['rival_tournaments'] = rival_tournaments

        if not effective_team and not is_admin:
            context['no_teams_assigned'] = True
            return context

        stats_service = StatsService(
            user=self.request.user,
            team_name=effective_team,
            seasons=selected_seasons if selected_seasons else None,
            tournaments=selected_tournaments if selected_tournaments else None,
        )

        context['available_seasons'] = stats_service.get_available_seasons()
        context['available_tournaments'] = stats_service.get_available_tournaments()

        # Determinar rival: override manual o próximo partido
        rival_name = rival_override
        next_match = None
        if not rival_name:
            today = timezone.localdate()
            teams_upper = set((effective_team or '').strip().upper() for _ in [1])
            if effective_team:
                for obj in Team.objects.filter(name__iexact=effective_team).exclude(alias__isnull=True).exclude(alias=''):
                    teams_upper.add(obj.alias.strip().upper())
            if teams_upper:
                nq = Q()
                for t in teams_upper:
                    nq |= Q(home_team__iexact=t) | Q(away_team__iexact=t)
                next_match = (
                    Match.objects.filter(nq, match_date__gte=today)
                    .order_by('match_date', 'match_time')
                    .select_related('tournament')
                    .first()
                )
            if next_match:
                if next_match.home_team.upper() in teams_upper:
                    rival_name = next_match.away_team
                else:
                    rival_name = next_match.home_team

        context['next_match'] = next_match
        context['rival_name'] = rival_name

        # Todos los rivales conocidos (para selector manual)
        all_rivals = _get_known_rivals(effective_team)
        context['all_rivals'] = all_rivals

        # Torneos disponibles del rival
        rival_available_tournaments = stats_service.get_rival_available_tournaments(rival_name) if rival_name else []
        context['rival_available_tournaments'] = rival_available_tournaments

        # Estadísticas de ambos equipos
        my_stats = stats_service.get_my_team_comparison_stats()
        rival_stats = stats_service.get_rival_aggregate_stats(rival_name, tournament_names=rival_tournaments if rival_tournaments else None) if rival_name else {'no_data': True}
        rival_detail = stats_service.get_rival_detail_stats(rival_name, tournament_names=rival_tournaments if rival_tournaments else None) if rival_name else {'no_data': True}

        context['my_stats'] = my_stats
        context['rival_stats'] = rival_stats
        context['rival_detail'] = rival_detail
        context['my_stats_json'] = json.dumps(my_stats, default=str)
        context['rival_stats_json'] = json.dumps(rival_stats, default=str)
        context['rival_detail_json'] = json.dumps(rival_detail, default=str)

        return context


def _get_known_rivals(team_name: str):
    """Retorna la lista de equipos únicos que han jugado contra team_name."""
    if not team_name:
        return []
    upper = team_name.strip().upper()
    rivals = set()
    for m in Match.objects.filter(
        Q(home_team__iexact=upper) | Q(away_team__iexact=upper)
    ).values('home_team', 'away_team'):
        home = (m['home_team'] or '').strip().upper()
        away = (m['away_team'] or '').strip().upper()
        if home != upper:
            rivals.add(home.title())
        if away != upper:
            rivals.add(away.title())
    return sorted(rivals)


# --- API Endpoints para datos dinámicos ---

class DashboardAPIView(DashboardAccessMixin, View):
    """API JSON para datos del dashboard (para AJAX/fetch)."""

    def get(self, request, *args, **kwargs):
        action = kwargs.get('action', 'summary')

        selected_seasons = request.GET.getlist('season', [])
        selected_team = request.GET.get('team', None)

        stats_service = StatsService(
            user=request.user,
            team_name=selected_team,
            seasons=selected_seasons if selected_seasons else None
        )

        if action == 'summary':
            data = stats_service.get_summary_stats()
        elif action == 'recent':
            limit = int(request.GET.get('limit', 5))
            data = stats_service.get_recent_matches(limit=limit)
        elif action == 'plays':
            data = stats_service.get_plays_distribution()
        elif action == 'trend':
            n = int(request.GET.get('n', 10))
            data = stats_service.get_trend_data(last_n_matches=n)
        elif action == 'zones':
            data = stats_service.get_zone_heatmap_data()
        elif action == 'match':
            match_id = int(request.GET.get('match_id', 0))
            if match_id:
                data = stats_service.get_match_detailed_stats(match_id)
            else:
                data = {'error': 'match_id required'}
        elif action == 'compare':
            match_ids = request.GET.getlist('match_id', [])
            match_ids = [int(mid) for mid in match_ids if mid.isdigit()]
            data = stats_service.compare_matches(match_ids)
        elif action == 'seasons':
            data = {'seasons': stats_service.get_available_seasons()}
        else:
            data = {'error': 'Unknown action'}

        return JsonResponse(data, safe=False)

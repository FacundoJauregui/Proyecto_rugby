import os
import pandas as pd
from openpyxl import load_workbook
import warnings
from datetime import timedelta

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# Carpetas
carpeta_entrada = 'excel_longo'
carpeta_salida = 'excel_procesado'
archivo_bd = 'BD_longo.xlsx'
os.makedirs(carpeta_salida, exist_ok=True)

# Mapeo de categorías a grupos comunes
mapeo_grupos = {
    'P OUR': 'POSESION',
    'P OPP': 'POSESION',
    'SALIDA OUR': 'SALIDAS',
    'SALIDA OPP': 'SALIDAS',
    'SCRUM OUR': 'SCRUMS',
    'SCRUM OPP': 'SCRUMS',
    'LINE OUR': 'LINE',
    'LINE OPP': 'LINE',
    'MAUL OUR': 'MAULS',
    'MAUL OPP': 'MAULS',
    'JUEGO': 'SECUENCIA',
    'PAUSA': 'OUT',
    'AMARILLA OUR': 'TARJETAS',
    'AMARILLA OPP': 'TARJETAS',
    'ROJA OUR': 'TARJETAS',
    'ROJA OPP': 'TARJETAS',
    'RUCKS GANADOS OPP': 'RUCKS',
    'RUCKS GANADOS OUR': 'RUCKS',
    'RUCKS PERDIDOS OUR': 'RUCKS_PERDIDO',
    'RUCKS PERDIDOS OPP': 'RUCKS_PERDIDO',
    'PENAL/FK OUR': 'PENALES_FK',
    'PENAL/FK OPP': 'PENALES_FK',
    'DROP OUR': 'DROPS',
    'DROP OPP': 'DROPS',
    'GOAL OUR': 'GOALS',
    'GOAL OPP': 'GOALS',
    'CONV OUR': 'CONVERSIONES',
    'CONV OPP': 'CONVERSIONES',
    'TRY OUR': 'TRIES',
    'TRY OPP': 'TRIES',
    'TRY P. OUR': 'TRIES',
    'TRY P. OPP': 'TRIES',
    'TRY DESDE OUR': 'TRIES_DESDE',  
    'TRY DESDE OPP': 'TRIES_DESDE',  
    'TO OPP': 'PELOTA_PERDIDA',
    'TO OUR': 'PELOTA_PERDIDA',
    'BREAKLINE OUR': 'BREAKLINE',
    'BREAKLINE OPP': 'BREAKLINE',
    'KICK OUR': 'KICKS',
    'KICK OPP': 'KICKS',
    'KICK MALO OUR': 'KICKS_MALO',
    'KICK MALO OPP': 'KICKS_MALO',
    'GOAL KICK OUR': 'GOAL_ERRADOS',
    'GOAL KICK OPP': 'GOAL_ERRADOS', 
    'KILLER INSTINT OUR': 'KILLER_INSTINT',
    'KILLER INSTINT OPP': 'KILLER_INSTINT',
    'PENAL OUR': 'PENALES_CONCEDIDOS',
    'PENAL OPP': 'PENALES_CONCEDIDOS',
    'CANCHA OUR': 'CANCHA',
    'CANCHA OPP': 'CANCHA',
    'TACKLE OUR': 'TACKLES',
    'TACKLE OPP': 'TACKLES',
    'PASE OUR': 'PASE',
    'PASE OPP': 'PASE',
    'CARRY OUR': 'CARRIES',
    'CARRY OPP': 'CARRIES',
    'OUR (IG a 50)': 'ZONA_DFF',
    'OPP(IG a 50)': 'ZONA_DFF',
    'OPP(50 a IG)': 'ZONA_ATT',
    'OUR(50 a IG)': 'ZONA_ATT',
    'LANZA OUR': 'LANZAMIENTOS',
    'LANZA OPP': 'LANZAMIENTOS',
    'HLG': 'HIGLIGTHS',
    'MEDICO': 'MEDICO',
    'Sustituciones': 'SUSTITUCIONES',
}

# Validar nombre de hoja
def nombre_valido(nombre):
    return nombre.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')[:31]

# Normalizar columnas
def normalizar_columnas(cols):
    return [
        str(col).strip().upper()
        .replace('\n', ' ')
        .replace('\r', ' ')
        .replace('Á', 'A').replace('É', 'E').replace('Í', 'I')
        .replace('Ó', 'O').replace('Ú', 'U').replace('Ü', 'U').replace('Ñ', 'N')
        for col in cols
    ]

# Extraer metadatos
def extraer_metadatos(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        return {
            'fecha_partido': ws['B3'].value or '',
            'torneo': f"{ws['B4'].value or ''} {ws['B5'].value or ''}".strip(),
            'equipo_local': ws['B6'].value or '',
            'equipo_visitante': ws['B7'].value or '',
            'arbitro': ws['B8'].value or '',
            'ficha': ws['B9'].value or '',
            'resultado': ws['B10'].value or '',
        }
    except Exception as e:
        print(f"⚠️ Error al leer metadatos: {e}")
        return {}

# Convertir a timedelta
def tiempo_a_timedelta(valor):
    if pd.isna(valor) or valor == '' or valor is None:
        return timedelta()
    try:
        if isinstance(valor, str):
            partes = list(map(int, valor.split(':')))
            if len(partes) == 2:
                m, s = partes
                return timedelta(minutes=m, seconds=s)
            elif len(partes) == 3:
                h, m, s = partes
                return timedelta(hours=h, minutes=m, seconds=s)
            else:
                return timedelta()
        elif isinstance(valor, timedelta):
            return valor
        else:
            # Para objetos time o similares
            return timedelta(
                hours=getattr(valor, 'hour', 0),
                minutes=getattr(valor, 'minute', 0),
                seconds=getattr(valor, 'second', 0)
            )
    except:
        return timedelta()

# Extraer secciones
def extraer_secciones_con_metadatos(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"❌ No se pudo abrir el archivo: {e}")
        return {}

    meta = extraer_metadatos(ruta_archivo)
    if not meta:
        return {}

    datos = {}
    current_section = None
    headers = None
    data_rows = []

    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        first_cell = row_list[0] if row_list else None

        # Detectar nueva sección
        if first_cell in mapeo_grupos:
            if current_section and headers and data_rows:
                df = pd.DataFrame(data_rows, columns=headers)
                equipo = meta['equipo_local'] if 'OUR' in current_section else meta['equipo_visitante'] if 'OPP' in current_section else ''
                df['Equipo'] = equipo
                df['Torneo'] = meta['torneo']
                df['Ficha'] = meta['ficha']
                df['Resultado'] = meta['resultado']
                df['Arbitro'] = meta['arbitro']

                grupo = mapeo_grupos[current_section]
                if grupo not in datos:
                    datos[grupo] = []
                datos[grupo].append(df)

            current_section = first_cell
            headers = None
            data_rows = []
            continue

        # Detectar encabezados
        if current_section and not headers:
            raw_row = [str(c) if c is not None else '' for c in row_list]
            if 'Tiempo' in raw_row or 'Evento' in raw_row or 'tiempo' in [r.lower() for r in raw_row]:
                headers = normalizar_columnas(row_list)
                seen = []
                unique_headers = []
                for h in headers:
                    if h in seen:
                        unique_headers.append(f"{h}_DUP{seen.count(h)}")
                    else:
                        unique_headers.append(h)
                    seen.append(h)
                headers = unique_headers
                continue

        # Recolectar filas de datos
        if current_section and headers:
            clean_row = [cell if cell is not None else '' for cell in row_list]
            while len(clean_row) < len(headers):
                clean_row.append('')
            if any(str(cell).strip() != '' for cell in clean_row):
                data_rows.append(clean_row)

    # Guardar última sección
    if current_section and headers and data_rows:
        df = pd.DataFrame(data_rows, columns=headers)
        equipo = meta['equipo_local'] if 'OUR' in current_section else meta['equipo_visitante'] if 'OPP' in current_section else ''
        df['Equipo'] = equipo
        df['Torneo'] = meta['torneo']
        df['Ficha'] = meta['ficha']
        df['Resultado'] = meta['resultado']
        df['Arbitro'] = meta['arbitro']
        grupo = mapeo_grupos[current_section]
        if grupo not in datos:
            datos[grupo] = []
        datos[grupo].append(df)

    # Combinar por grupo
    resultado = {}
    for grupo, dfs in datos.items():
        if not dfs:
            continue
        try:
            # Alinear columnas
            all_cols = sorted(set(col for df in dfs for col in df.columns))
            aligned_dfs = []
            for df in dfs:
                for col in all_cols:
                    if col not in df.columns:
                        df[col] = ''
                aligned_dfs.append(df[all_cols])
            combined = pd.concat(aligned_dfs, ignore_index=True)
            combined.dropna(how='all', inplace=True)

            # Eliminar columnas innecesarias
            cols_to_drop = [col for col in combined.columns if col.startswith('Col_') or combined[col].astype(str).str.strip().eq('').all()]
            combined.drop(columns=cols_to_drop, inplace=True, errors='ignore')

            # ✅ Cálculo de resultado_tiempo en POSESION, OUT y SECUENCIA
            if grupo in ['POSESION', 'OUT', 'SECUENCIA']:
                if 'TIEMPO' in combined.columns and 'FIN' in combined.columns:
                    def calcular_diferencia(row):
                        t1 = tiempo_a_timedelta(row['TIEMPO'])
                        t2 = tiempo_a_timedelta(row['FIN'])
                        diff = t2 - t1
                        if diff.total_seconds() < 0:
                            diff = timedelta()  # Evitar tiempos negativos
                        total_seg = int(diff.total_seconds())
                        mins = total_seg // 60
                        secs = total_seg % 60
                        return f"{mins:02d}:{secs:02d}"
                    combined['resultado_tiempo'] = combined.apply(calcular_diferencia, axis=1)

            if not combined.empty:
                resultado[grupo] = combined
        except Exception as e:
            print(f"❌ Error en grupo '{grupo}': {e}")
            continue

    return resultado

# === PASO 1: Procesar archivos ===
print("🚀 Iniciando procesamiento de archivos...")
for archivo in os.listdir(carpeta_entrada):
    if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
        entrada = os.path.join(carpeta_entrada, archivo)
        salida = os.path.join(carpeta_salida, f'procesado_{archivo}')
        print(f"  Procesando: {archivo}")
        try:
            secciones = extraer_secciones_con_metadatos(entrada)
            if not secciones:
                print(f"    ⚠️  Sin datos: {archivo}")
                continue
            with pd.ExcelWriter(salida, engine='openpyxl') as writer:
                for grupo, df in secciones.items():
                    if not df.empty:
                        sheet = nombre_valido(grupo)
                        df.to_excel(writer, sheet_name=sheet, index=False)
            print(f"  ✅ Guardado: {salida}")
        except Exception as e:
            print(f"  ❌ Error procesando {archivo}: {e}")

# === PASO 2: Consolidar en BD_longo.xlsx ===
print("\n📊 Consolidando en BD_longo.xlsx...")
archivos_proc = [f for f in os.listdir(carpeta_salida) if f.startswith('procesado_') and f.endswith('.xlsx')]

if not archivos_proc:
    print("⚠️ No hay archivos procesados.")
else:
    datos_bd = {}
    for arch in archivos_proc:
        ruta = os.path.join(carpeta_salida, arch)
        print(f"  Leyendo: {arch}")
        try:
            xl = pd.ExcelFile(ruta)
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                if df.empty:
                    continue
                df['Archivo_Origen'] = arch
                key = nombre_valido(sheet)
                if key not in datos_bd:
                    datos_bd[key] = []
                datos_bd[key].append(df)
        except Exception as e:
            print(f"  ❌ Error leyendo {arch}: {e}")

    try:
        with pd.ExcelWriter(archivo_bd, engine='openpyxl') as writer:
            for sheet_name, dfs in datos_bd.items():
                final_df = pd.concat(dfs, ignore_index=True)
                final_df.dropna(how='all', inplace=True)
                if not final_df.empty:
                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"✅ Consolidado: '{archivo_bd}'")
    except Exception as e:
        print(f"❌ Error al guardar '{archivo_bd}': {e}")

print("✅ Proceso completado.")